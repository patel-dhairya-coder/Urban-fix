# admin_dashboard/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db import models # Keep this import for Q objects
from django.db.models import Count, Sum, Case, When, IntegerField
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import timedelta
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from user.models import Complaint, Contractor  # Import Contractor model
from django.contrib.auth.models import User

# --- Helper function to check if user is staff/superuser ---
# ... (is_admin function remains the same) ...
def is_admin(user):
    return user.is_staff or user.is_superuser

# --- Custom Admin Login ---
# ... (admin_login view remains the same) ...
def admin_login(request):
    if request.user.is_authenticated and is_admin(request.user):
        return redirect('admin_dashboard:home')

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None and is_admin(user):
            login(request, user)
            messages.success(request, f"Welcome, Admin {username}!")
            return redirect('admin_dashboard:home')
        else:
            messages.error(request, "Invalid credentials or not an admin user.")
    return render(request, "admin_dashboard/login.html")

# --- Admin Logout ---
# ... (admin_logout view remains the same) ...
@login_required
def admin_logout(request):
    if not is_admin(request.user):
        return redirect('user:home')
    logout(request)
    messages.success(request, "You have been logged out from the admin panel.")
    return redirect('admin_dashboard:login')

# --- Admin Dashboard Home (Statistics & Analytics) ---
# ... (dashboard_home view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def dashboard_home(request):
    # Total counts
    total_complaints = Complaint.objects.count()
    status_counts = Complaint.objects.values('status').annotate(total=Count('status'))
    status_dict = {item['status']: item['total'] for item in status_counts}

    total_registered_users = User.objects.filter(is_staff=False, is_superuser=False).count()

    # Complaints per category
    complaints_per_category = Complaint.objects.values('category').annotate(count=Count('category')).order_by('-count')

    # Monthly complaints trend (last 6 months)
    today = timezone.now()
    six_months_ago = today - timedelta(days=180)
    monthly_complaints = Complaint.objects.filter(submitted_at__gte=six_months_ago) \
                                        .annotate(month=TruncMonth('submitted_at')) \
                                        .values('month') \
                                        .annotate(count=Count('report_id')) \
                                        .order_by('month')

    monthly_labels = [m['month'].strftime("%b %Y") for m in monthly_complaints]
    monthly_data = [m['count'] for m in monthly_complaints]

    # --- Contractor KPIs for Dashboard Home ---
    total_contractors = Contractor.objects.count()
    active_contractors = Contractor.objects.filter(is_active=True).count()
    completed_tasks = Complaint.objects.filter(status='resolved', assigned_to__isnull=False).count()
    pending_tasks = Complaint.objects.filter(status__in=['pending', 'in_progress'], assigned_to__isnull=False).count()


    context = {
        'total_complaints': total_complaints,
        'status_counts': status_dict,
        'total_registered_users': total_registered_users,
        'complaints_per_category': complaints_per_category,
        'monthly_labels': monthly_labels,
        'monthly_data': monthly_data,
        # New Contractor KPIs
        'total_contractors': total_contractors,
        'active_contractors': active_contractors,
        'completed_tasks': completed_tasks,
        'pending_tasks': pending_tasks,
    }
    return render(request, "admin_dashboard/dashboard.html", context)

# --- Complaint Management: List with Filters ---
# ... (complaint_list view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def complaint_list(request):
    complaints = Complaint.objects.all().select_related('user', 'assigned_to') # Also prefetch assigned_to
    # ... (filtering and search logic remains the same) ...

    # Filtering
    category_filter = request.GET.get('category')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search')
    contractor_filter = request.GET.get('contractor') # New filter for contractors

    if category_filter:
        complaints = complaints.filter(category=category_filter)
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    if search_query:
        complaints = complaints.filter(
            models.Q(report_id__icontains=search_query) |
            models.Q(category__icontains=search_query) |
            models.Q(user__username__icontains=search_query)
        )
    if contractor_filter:
        if contractor_filter == 'unassigned':
            complaints = complaints.filter(assigned_to__isnull=True)
        elif contractor_filter == 'assigned':
            complaints = complaints.filter(assigned_to__isnull=False)
        else:
            complaints = complaints.filter(assigned_to__id=contractor_filter)


    # Get available categories and statuses for filters
    categories = Complaint.CATEGORY_CHOICES
    statuses = Complaint.STATUS_CHOICES
    contractors = Contractor.objects.filter(is_active=True).order_by('name') # Active contractors for filter

    context = {
        'complaints': complaints,
        'categories': categories,
        'statuses': statuses,
        'contractors': contractors, # Pass contractors for filter dropdown
        'selected_category': category_filter,
        'selected_status': status_filter,
        'selected_contractor': contractor_filter,
        'search_query': search_query,
    }
    return render(request, "admin_dashboard/complaint_list.html", context)

# --- Complaint Management: Detail & Contractor Assignment ---
# ... (complaint_detail view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def complaint_detail(request, report_id):
    complaint = get_object_or_404(Complaint, report_id=report_id)
    active_contractors = Contractor.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        assigned_contractor_id = request.POST.get('assigned_to')

        # Handle contractor assignment
        if assigned_contractor_id:
            if assigned_contractor_id == 'none': # Option to unassign
                complaint.assigned_to = None
                complaint.assigned_at = None
                messages.info(request, "Contractor unassigned.")

                # --- NEW LOGIC: Revert status to 'pending' if it was 'in_progress' ---
                if complaint.status == 'in_progress':
                    complaint.status = 'pending'
                    messages.info(request, "Complaint status reverted to 'Pending'.")
                # --- END NEW LOGIC ---

            else:
                contractor = get_object_or_404(Contractor, id=assigned_contractor_id)
                if complaint.assigned_to != contractor: # Only update if changed
                    complaint.assigned_to = contractor
                    messages.success(request, f"Complaint {report_id} assigned to {contractor.name}.")
                    
                    # --- NEW LOGIC: Update status to 'in_progress' if it was 'pending' ---
                    if complaint.status == 'pending':
                        complaint.status = 'in_progress'
                        messages.info(request, "Complaint status updated to 'In Progress'.")
                    # --- END NEW LOGIC ---
        
        elif 'assigned_to' in request.POST and complaint.assigned_to: # If assigned_to field was present but value was empty/not 'none'
            complaint.assigned_to = None
            complaint.assigned_at = None
            messages.info(request, "Contractor unassigned (no selection made).")
            # --- NEW LOGIC: Revert status to 'pending' if it was 'in_progress' ---
            if complaint.status == 'in_progress':
                complaint.status = 'pending'
                messages.info(request, "Complaint status reverted to 'Pending'.")
            # --- END NEW LOGIC ---


        complaint.save() # Save changes made to assigned_to AND status
        return redirect('admin_dashboard:complaint_detail', report_id=report_id)
    else:
        # If it's a GET request, just render the page
        pass

    context = {
        'complaint': complaint,
        'active_contractors': active_contractors,
    }
    return render(request, "admin_dashboard/complaint_detail.html", context)

# --- Complaint Delete View ---
# ... (delete_complaint view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def delete_complaint(request, report_id):
    complaint = get_object_or_404(Complaint, report_id=report_id)
    if request.method == 'POST':
        try:
            complaint_id_temp = complaint.report_id
            complaint.delete()
            messages.success(request, f"Complaint '{complaint_id_temp}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting complaint: {e}")
    # Redirect to the complaint list (even if it was a GET request, just send them back)
    return redirect('admin_dashboard:complaint_list')


# --- User Management (remains the same) ---
# ... (user_list view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def user_list(request):
    users = User.objects.filter(is_staff=False, is_superuser=False).annotate(
        complaint_count=Count('complaint')
    ).order_by('username')
    context = { 'users': users, }
    return render(request, "admin_dashboard/user_list.html", context)

# ... (deactivate_user view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def deactivate_user(request, user_id):
    user_to_deactivate = get_object_or_404(User, id=user_id, is_staff=False, is_superuser=False)
    if request.method == 'POST':
        user_to_deactivate.is_active = not user_to_deactivate.is_active
        user_to_deactivate.save()
        status = "deactivated" if not user_to_deactivate.is_active else "activated"
        messages.success(request, f"User '{user_to_deactivate.username}' has been {status}.")
    return redirect('admin_dashboard:user_list')

# --- Reports & Export (remain the same) ---
# ... (export_complaints view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def export_complaints(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="urbanfix_complaints.xlsx"'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Complaints"
    headers = ["Report ID", "User", "Category", "Location", "Description", "Submitted At", "Status", "Assigned To"]
    worksheet.append(headers)
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f"{col_letter}1"].font = Font(bold=True)
        worksheet[f"{col_letter}1"].alignment = Alignment(horizontal="center")
        worksheet.column_dimensions[col_letter].width = 18
    complaints = Complaint.objects.all().select_related('user', 'assigned_to').order_by('-submitted_at')
    for complaint in complaints:
        row = [
            complaint.report_id,
            complaint.user.username,
            complaint.get_category_display(),
            complaint.location,
            complaint.description,
            complaint.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
            complaint.get_status_display(),
            complaint.assigned_to.name if complaint.assigned_to else "Unassigned", # New field for export
        ]
        worksheet.append(row)
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    workbook.save(response)
    return response

# ... (monthly_summary_report view remains the same) ...
@user_passes_test(is_admin, login_url='/admin-panel/login/')
def monthly_summary_report(request):
    today = timezone.now()
    one_year_ago = today - timedelta(days=365)
    monthly_data = Complaint.objects.filter(submitted_at__gte=one_year_ago) \
        .annotate(month=TruncMonth('submitted_at')) \
        .values('month') \
        .annotate(
            total=Count('id'),
            pending=Sum(Case(When(status='pending', then=1), default=0, output_field=IntegerField())),
            in_progress=Sum(Case(When(status='in_progress', then=1), default=0, output_field=IntegerField())),
            resolved=Sum(Case(When(status='resolved', then=1), default=0, output_field=IntegerField())),
            rejected=Sum(Case(When(status='rejected', then=1), default=0, output_field=IntegerField())),
        ).order_by('month')
    context = { 'monthly_data': monthly_data, }
    return render(request, "admin_dashboard/monthly_summary.html", context)


# --- MODIFIED: Contractor Management Views ---

@user_passes_test(is_admin, login_url='/admin-panel/login/')
def contractor_list(request):
    # ... (contractor_list view remains the same) ...
    contractors = Contractor.objects.annotate(
        total_assigned_tasks=Count('assigned_complaints')
    ).order_by('name')

    context = {
        'contractors': contractors,
    }
    return render(request, "admin_dashboard/contractor_list.html", context)

@user_passes_test(is_admin, login_url='/admin-panel/login/')
def add_contractor(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        contact_number = request.POST.get('contact_number')
        email = request.POST.get('email')
        password = request.POST.get('password') # <-- NEW
        specialization = request.POST.get('specialization')
        area_assigned = request.POST.get('area_assigned')
        is_active = request.POST.get('is_active') == 'on' 

        # --- MODIFIED: Added password to check ---
        if not all([name, email, specialization, password]):
            messages.error(request, "Please fill in all required fields (Name, Email, Password, Specialization).")
            # Re-render form with existing data if validation fails
            context = {
                'contractor': {
                    'name': name, 'contact_number': contact_number, 'email': email,
                    'specialization': specialization, 'area_assigned': area_assigned,
                    'is_active': is_active
                },
                'specialization_choices': Contractor.SPECIALIZATION_CHOICES
            }
            return render(request, "admin_dashboard/contractor_form.html", context)

        try:
            # --- MODIFIED: Added password to create ---
            Contractor.objects.create(
                name=name,
                contact_number=contact_number,
                email=email,
                password=password, # The model's save() method will hash it
                specialization=specialization,
                area_assigned=area_assigned,
                is_active=is_active
            )
            messages.success(request, f"Contractor '{name}' added successfully!")
            return redirect('admin_dashboard:contractor_list')
        except Exception as e:
            messages.error(request, f"Error adding contractor: {e}")
            context = {
                'contractor': {
                    'name': name, 'contact_number': contact_number, 'email': email,
                    'specialization': specialization, 'area_assigned': area_assigned,
                    'is_active': is_active
                },
                'specialization_choices': Contractor.SPECIALIZATION_CHOICES
            }
            return render(request, "admin_dashboard/contractor_form.html", context)
    else:
        context = {
            'specialization_choices': Contractor.SPECIALIZATION_CHOICES,
            'contractor': {'is_active': True} # Default active for new contractors
        }
    return render(request, "admin_dashboard/contractor_form.html", context)

@user_passes_test(is_admin, login_url='/admin-panel/login/')
def edit_contractor(request, contractor_id):
    contractor = get_object_or_404(Contractor, id=contractor_id)

    if request.method == 'POST':
        contractor.name = request.POST.get('name')
        contractor.contact_number = request.POST.get('contact_number')
        contractor.email = request.POST.get('email')
        contractor.specialization = request.POST.get('specialization')
        contractor.area_assigned = request.POST.get('area_assigned')
        contractor.is_active = request.POST.get('is_active') == 'on'
        
        # --- NEW: Handle password change ---
        password = request.POST.get('password')
        if password:
            contractor.password = password # The model's save() method will hash it
        # --- END NEW ---

        if not all([contractor.name, contractor.email, contractor.specialization]):
            messages.error(request, "Please fill in all required fields (Name, Email, Specialization).")
            # Re-render form with existing data if validation fails
            context = {
                'contractor': contractor,
                'specialization_choices': Contractor.SPECIALIZATION_CHOICES
            }
            return render(request, "admin_dashboard/contractor_form.html", context)

        try:
            contractor.save()
            messages.success(request, f"Contractor '{contractor.name}' updated successfully!")
            return redirect('admin_dashboard:contractor_list')
        except Exception as e:
            messages.error(request, f"Error updating contractor: {e}")
            context = {
                'contractor': contractor,
                'specialization_choices': Contractor.SPECIALIZATION_CHOICES
            }
            return render(request, "admin_dashboard/contractor_form.html", context)
    else:
        context = {
            'contractor': contractor,
            'specialization_choices': Contractor.SPECIALIZATION_CHOICES,
        }
    return render(request, "admin_dashboard/contractor_form.html", context)

@user_passes_test(is_admin, login_url='/admin-panel/login/')
def delete_contractor(request, contractor_id):
    # ... (delete_contractor view remains the same) ...
    contractor = get_object_or_404(Contractor, id=contractor_id)
    if request.method == 'POST':
        try:
            # Check if there are any complaints currently assigned to this contractor
            if Complaint.objects.filter(assigned_to=contractor).exists():
                messages.error(request, f"Cannot delete contractor '{contractor.name}' because there are complaints assigned to them. Please reassign them first.")
                return redirect('admin_dashboard:contractor_list')

            contractor.delete()
            messages.success(request, f"Contractor '{contractor.name}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting contractor: {e}")
    return redirect('admin_dashboard:contractor_list')


@user_passes_test(is_admin, login_url='/admin-panel/login/')
def contractor_analytics(request):
    # ... (contractor_analytics view remains the same) ...
    total_contractors = Contractor.objects.count()
    active_contractors = Contractor.objects.filter(is_active=True).count()

    # Complaints assigned per contractor
    contractors_assigned_counts = Contractor.objects.annotate(
        assigned_count=Count('assigned_complaints')
    ).order_by('-assigned_count')

    assigned_labels = [c.name for c in contractors_assigned_counts]
    assigned_data = [c.assigned_count for c in contractors_assigned_counts]

    # Task distribution by status for assigned complaints
    task_distribution_assigned = Complaint.objects.filter(assigned_to__isnull=False) \
        .values('status') \
        .annotate(total=Count('status'))

    task_dist_labels = [item['status'].replace('_', ' ').title() for item in task_distribution_assigned]
    task_dist_data = [item['total'] for item in task_distribution_assigned]

    # Calculate completed vs pending tasks (for assigned tasks only)
    assigned_resolved_tasks = Complaint.objects.filter(assigned_to__isnull=False, status='resolved').count()
    assigned_pending_tasks = Complaint.objects.filter(assigned_to__isnull=False, status__in=['pending', 'in_progress']).count()

    context = {
        'total_contractors': total_contractors,
        'active_contractors': active_contractors,
        'assigned_resolved_tasks': assigned_resolved_tasks,
        'assigned_pending_tasks': assigned_pending_tasks,
        'assigned_labels': assigned_labels,
        'assigned_data': assigned_data,
        'task_dist_labels': task_dist_labels,
        'task_dist_data': task_dist_data,
    }
    return render(request, "admin_dashboard/contractor_analytics.html", context)

def home(request):
    return render(request, "user/home.html")